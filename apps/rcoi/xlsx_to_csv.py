import csv
import glob
import logging
import os
import re
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

logger = logging.getLogger(__name__)
logging.basicConfig(
    format='%(asctime)s [%(name)s:%(lineno)s] %(levelname)s - %(message)s',
    level=logging.DEBUG)
logging.getLogger('urllib3').setLevel(logging.WARNING)


def get_filename_from_url(url):
    """
    Format filename from parts of URL

    :type url: str
    :param url: file url
    :return: formatted filename
    :rtype: str
    """
    url_parts = url.split('/')
    filename = '{}-{}-{}'.format(url_parts[-4], url_parts[-3], url_parts[-1])
    return filename


def get_files_info(url):
    """
    Get remote file headers from URL

    :type url: str
    :param url: file url
    :return: files headers
    :rtype: dict
    """
    logger.debug('get file links: %s', url)
    headers = {
        'content-type': 'application/x-www-form-urlencoded',
        'cache-control': 'no-cache'
    }
    fmt_dt = '%a, %d %b %Y %H:%M:%S %Z'
    url_base = '/'.join(url.split('/')[:3])
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'lxml')
    content = soup.select('span[data-class="info"]')
    blocks = [(block.attrs.get('data-id'), block.attrs.get('data-ident')) for block in content]

    links = []
    for block in blocks:
        payload = 'id={}&data={}&val=1'.format(*block)
        r2 = requests.request('POST', url, data=payload, headers=headers)
        soup2 = BeautifulSoup(r2.text, 'lxml')
        content2 = soup2.select('p a')
        links += [url_base + a.attrs.get('href') for a in content2 if 'rab' in a.attrs.get('href')]

    files_info = []
    for url in links:
        local_filename = get_filename_from_url(url)
        r3 = requests.head(url)
        lmd = datetime.strptime(r3.headers['Last-Modified'], fmt_dt)
        file = {
            'name': local_filename,
            'url': url,
            'size': r3.headers['Content-Length'],
            'last_modified': lmd
        }
        files_info.append(file)
    return files_info


def download_file(url, path):
    """
    Download file from URL

    :type url: str
    :param url: file url
    :type path: str
    :param path: local path
    """
    local_filename = get_filename_from_url(url)
    logger.debug('download file: %s', local_filename)
    fs = requests.get(url, stream=True)
    with open(os.path.join(path, local_filename), 'wb') as f:
        for chunk in fs.iter_content(chunk_size=1024):
            if chunk:
                f.write(chunk)


# replace quotation marks
s0 = re.compile(r'["„“”«»\'‘’]')
# remove spaces before punctuation marks
s1 = re.compile(r'\s+(?=[.,:;!?])')
# 1. (add space after punctuation marks |
# 2. add space before "№" |
# 3. replace whitespace chars with only one space)
s2 = re.compile(r'(?<=[.,:;!№?])(?![.,:;!№?\s])|(?<=\w)(?=№)|\s+')


def re_work(s):
    """
    Format string with precompiled regexp rules

    :type s: str
    :param s: string
    :return: formatted string
    :rtype: str
    """
    s = s2.sub(' ', s1.sub('', s0.sub(' ', str(s)))).strip()
    return s


def rename_org(value):
    """
    Replace parts of organization name

    :type value: str
    :param value: original name
    :return: modified name
    :rtype: str
    """
    value = value.replace('Федеральное государственное бюджетное общеобразовательное учреждение высшего образования',
                          'ФГБОУ ВО')
    value = value.replace('федеральное государственное бюджетное общеобразовательное учреждение высшего образования',
                          'ФГБОУ ВО')
    value = value.replace('Федеральное государственное бюджетное общеобразовательное учреждение',
                          'ФГБОУ')
    value = value.replace('Автономная некоммерческая образовательная организация',
                          'АНОО')
    value = value.replace('Автономная некомерческая организация высшего образования',
                          'АНОВО')
    value = value.replace('Автономная некоммерческая общеобразовательная организация',
                          'АНОО')
    value = value.replace('Государственное автономное общеобразовательное учреждение города Москвы',
                          'ГАОУ')
    value = value.replace('Государственное автономное общеобразовательное учреждение дополнительного профессионального образования города Москвы',  # noqa
                          'ГАОУ ДПО')
    value = value.replace('Государственное автономное общеобразовательное учреждение высшего образования города Москвы',
                          'ГАОУ ВО')
    value = value.replace('Государственное автономное профессиональное общеобразовательное учреждение города Москвы',
                          'ГАПОУ')
    value = value.replace('Государственное бюджетное профессиональное общеобразовательное учреждение города Москвы',
                          'ГБПОУ')
    value = value.replace('Государственное бюджетное профессиональное общеобразовательное учреждение (колледж) города Москвы',  # noqa
                          'ГБПОУ')
    value = value.replace('Государственное бюджетное профессиональное общеобразовательное учреждение г. Москвы',
                          'ГБПОУ')
    value = value.replace('ГБПОУ г. Москвы',
                          'ГБПОУ')
    value = value.replace('Государственное бюджетное общеобразовательное учреждение города Москвы',
                          'ГБОУ')
    value = value.replace('Государственное бюджетное общеобразовательноеучреждение города Москвы',
                          'ГБОУ')
    value = value.replace('Государственное бюджетное общеобразовательное учреждение',
                          'ГБОУ')
    value = value.replace('государственное бюджетное общеобразовательное учреждение',
                          'ГБОУ')
    value = value.replace('Государственное бюджетное учреждение города Москвы',
                          'ГБУ')
    value = value.replace('Государственное бюджетное учреждение средняя общеобразовательная школа',
                          'ГБУ СОШ')
    value = value.replace('Государственное казенное общеобразовательное учреждение города Москвы',
                          'ГКОУ')
    value = value.replace('Муниципальное автономное общеобразовательное учреждение',
                          'МАОУ')
    value = value.replace('Негосударственная общеобразовательная организация частное учреждение',
                          'НООЧУ')
    value = value.replace('Негосударственное образовательное частное учреждение',
                          'НОЧУ')
    value = value.replace('Негосударственное образовательное частное учреждения',
                          'НОЧУ')
    value = value.replace('Негосударственное некоммерческое общеобразовательное учреждение',
                          'ННОУ')
    value = value.replace('Негосударственное общеобразовательное учреждение',
                          'НОУ')
    value = value.replace('Негосударственное общеобразовательное частное учреждение',
                          'НОЧУ')
    value = value.replace('Негосударственное частное учреждение общеобразовательная организация',
                          'НЧУОО')
    value = value.replace('Некоммерческое образовательное частное учреждение',
                          'НОЧУ')
    value = value.replace('Общеобразовательная автономная некоммерческая организация',
                          'ОАНО')
    value = value.replace('Автономная некоммерческая организация',
                          'АНО')
    value = value.replace('Общеобразовательное частное учреждение',
                          'ОЧУ')
    value = value.replace('Образовательное частное учреждение',
                          'ОЧУ')
    value = value.replace('Общеобразовательная организация частное учреждение',
                          'ООЧУ')
    value = value.replace('Общеобщеобразовательная автономная некоммерческая организация',
                          'ОАНО')
    value = value.replace('средняя общеобразовательная школа',
                          'СОШ')
    value = value.replace('Средняя общеобразовательная школа',
                          'СОШ')
    value = value.replace('Федеральное государственное автономное общеобразовательное учреждение',
                          'ФГАОУ')
    value = value.replace('федеральное государственное бюджетное общеобразовательное учреждение',
                          'ФГБОУ')
    value = value.replace('Федеральное государственное бюджетное общеобразовательное учреждение',
                          'ФГБОУ')
    value = value.replace('Федеральное государственное казенное общеобразовательное учреждение',
                          'ФГКОУ')
    value = value.replace('Федеральное государственное казённое общеобразовательное учреждение',
                          'ФГКОУ')
    value = value.replace('Частное общеобразовательное учреждение',
                          'ЧОУ')
    value = value.replace('Частное учреждение общеобразовательная организация',
                          'ЧУ ОО')
    value = value.replace('Частное учреждение Общеобразовательная организация',
                          'ЧУ ОО')
    value = value.replace('Частное учреждение средняя общеобразовательная школа',
                          'ЧУ СОШ')
    value = value.replace('Частное учреждение средняя общеобразовательное школа',
                          'ЧУ СОШ')
    value = value.replace('Частное учреждение Средняя общеобразовательная школа',
                          'ЧУ СОШ')
    value = value.replace('Частное учреждение',
                          'ЧУ')
    return value


def extract_date(header):
    """
    Extract date from string

    :type header: list
    :param header: list with text date
    :return: formatted text date
    :rtype: str
    """
    months = {
        'января': '01', 'февраля': '02', 'марта': '03',
        'апреля': '04', 'мая': '05', 'июня': '06',
        'июля': '07', 'августа': '08', 'сентября': '09',
        'октября': '10', 'ноября': '11', 'декабря': '12'
    }
    date = header[-4:-1]
    date[1] = months[date[1]]
    if len(date[0]) == 1:
        date[0] = '0' + date[0]
    return '-'.join(reversed(date))


def parse_xlsx(filename):
    """
    Process Excel file and convert it to list of rows

    :type filename: str
    :param filename: local xlsx file name
    :return: list of rows
    :rtype: list
    """
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    if len(tuple(ws.columns)) < 7:
        logger.debug('skip file %s: wrong number of columns', filename)
        return []
    data = tuple(ws.rows)
    row_count = len(data)
    header = data[0][0].value.split()
    if 'ГИА-11' in header:
        level = 11
    elif 'ГИА-9' in header:
        level = 9
    else:
        level = 'Другое'
    date = extract_date(header)
    filename = filename.split('/')[-1]
    logger.debug('parse file: %s (date %s, level %s, rows %s)', filename, date, level, row_count-2)
    l_data = []
    for row_num in range(2, row_count):
        row = data[row_num]
        l_row = [filename, date, level]
        # Пропускаем 1 ячейку с порядковым номером
        for cell_num in range(1, len(row)):
            cell = row[cell_num].value
            if cell:
                cell = re_work(cell)
                if ' образовательное учреждение' in cell:
                    cell = cell.replace(' образовательное учреждение',
                                        ' общеобразовательное учреждение')
                cell = rename_org(cell)
            l_row.append(cell)
        if l_row[5]:
            l_data.append(l_row)
    return l_data


def save_to_csv(csv_file):
    """
    Save list of rows to CSV in local file

    :type csv_file: str
    :param csv_file: local CSV file name
    """
    with open(csv_file, 'w+', newline='', encoding='utf-8') as fp:
        a = csv.writer(fp, delimiter=';')
        a.writerow(['datafile', 'date', 'level',
                    'ppe_code', 'ppe_name', 'ppe_addr',
                    'position', 'name', 'organisation'])

    xlsx = os.path.join(os.path.split(csv_file)[0], '*.xlsx')
    for name in glob.glob(xlsx):
        data = parse_xlsx(name)
        if not data:
            continue
        with open(csv_file, 'a+', newline='', encoding='utf-8') as fp:
            a = csv.writer(fp, delimiter=';')
            for line in data:
                a.writerow(line)


def save_to_stream(path):
    """
    Save list of rows to CSV in string buffer (memory file)

    :type path: str
    :param path: local directory with xlsx files
    :return: memory file
    :rtype: StringIO
    """
    from io import StringIO
    stream = StringIO()
    writer = csv.writer(stream, delimiter='\t', quotechar="'")
    writer.writerow(['datafile', 'date', 'level',
                     'ppe_code', 'ppe_name', 'ppe_addr',
                     'position', 'name', 'organisation'])

    xlsx = os.path.join(path, '*.xlsx')
    for name in glob.glob(xlsx):
        data = parse_xlsx(name)
        for line in data:
            writer.writerow(line)
    stream.seek(0)
    return stream


def main():
    path = 'data'
    if not os.path.exists(path):
        os.makedirs(path)
    csv_file = os.path.join(path, path + '.csv')
    urls = ['http://rcoi.mcko.ru/organizers/schedule/oge/?period=2',
            'http://rcoi.mcko.ru/organizers/schedule/ege/?period=2']
    files_info = [get_files_info(url) for url in urls]
    files_info = [url for url_list in files_info for url in url_list]
    [download_file(file['url'], path) for file in files_info]
    save_to_csv(csv_file)


if __name__ == '__main__':
    main()
