import csv
import logging
import re
from datetime import datetime
from io import StringIO
from multiprocessing.pool import ThreadPool
from pathlib import Path
from urllib.parse import urljoin

import requests
import stamina
from bs4 import BeautifulSoup, Comment
from openpyxl import load_workbook

logger = logging.getLogger(__name__)
logging.getLogger("urllib3").setLevel(logging.WARNING)

QUOTES = re.compile(r'[<>"„“”«»‘’\']')
"""regexp for quotation marks"""
PUNCTUATION = re.compile(r"\s+(?=[.,:;!?])")
"""regexp for spaces before punctuation marks"""
SPACES = re.compile(r"(?<=[.,:;!№?])(?![.,:;!№?\s])|(?<=\w)(?=№)|\s+")
"""regexp for (after punctuation marks | before "№" | whitespace chars)"""

COLUMN_HEADERS = [
    "datafile",
    "date",
    "level",
    "ppe_code",
    "ppe_name",
    "ppe_addr",
    "position",
    "name",
    "organisation",
]

client = requests.Session()


class InvalidFileError(Exception):
    """Invalid file error."""


@stamina.retry(on=requests.HTTPError, attempts=5)
def download_file(url: str, file_name: str, file_path: Path) -> None:
    """Download file from URL."""
    logger.debug("download file %s as %s", url, file_name)
    f = file_path.joinpath(file_name)
    if f.exists():
        logger.error(
            "There are more than 1 file for exam date: %s. SKIP download %s",
            file_name,
            url,
        )
        return
    with client.get(url, stream=True, timeout=5) as res:
        res.raise_for_status()
        with f.open("wb") as fp:
            for chunk in res.iter_content(chunk_size=8192):
                fp.write(chunk)


@stamina.retry(on=requests.HTTPError, attempts=5)
def prepare_file_info(url: str, date: str, level: str) -> dict:
    """Prepare info about single exam file from direct url."""
    logger.debug("get file info: %s", url)
    ext = Path(url).suffix
    local_filename = f"{date}__{level}__{ext}"

    res = client.head(url, stream=True, timeout=5)
    res.raise_for_status()
    last_modified = datetime.strptime(  # noqa: DTZ007
        res.headers["Last-Modified"],
        "%a, %d %b %Y %H:%M:%S %Z",
    )
    return {
        "name": local_filename,
        "url": url,
        "size": res.headers["Content-Length"],
        "last_modified": last_modified,
    }


@stamina.retry(on=requests.HTTPError, attempts=5)
def get_content_blocks_ids(url: str) -> list:
    """Get content blocks ids from exams page."""
    res = client.get(url, timeout=5)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml").select('span[data-class="info"]')
    return [
        (block.attrs.get("data-id"), block.attrs.get("data-ident")) for block in soup
    ]


@stamina.retry(on=requests.HTTPError, attempts=5)
def get_content_block_soup(url: str, data_id: str, data_ident: str) -> BeautifulSoup:
    """Get soup from specific exam content block."""
    res = client.post(
        url,
        data=f"id={data_id}&data={data_ident}&val=1",
        headers={
            "content-type": "application/x-www-form-urlencoded",
            "cache-control": "no-cache",
        },
        timeout=5,
    )
    res.raise_for_status()
    return BeautifulSoup(res.text, "lxml")


def extract_href(soup: BeautifulSoup) -> str | None:
    """Extract value of href attribute containing keywords "rab" or "работник"."""
    # remove comments
    for element in soup(string=lambda string: isinstance(string, Comment)):
        element.extract()
    for a in soup.select("p a"):
        # skip bad tags
        if not a.attrs:
            continue
        href = a.attrs.get("href", "")
        if "rab" in href.lower() or "работник" in href.lower():
            return href
    return None


def get_files_info(url: str) -> list:
    """Compose custom info about remote files."""
    logger.debug("get file links: %s", url)
    level = "0"
    if "/ege/" in url:
        level = "11"
    elif "/oge/" in url:
        level = "9"

    result = []
    for data_id, data_ident in get_content_blocks_ids(url):
        if not data_id or not data_ident:
            continue
        block_soup = get_content_block_soup(url, data_id, data_ident)
        if not (file_link := extract_href(block_soup)):
            continue
        date = f"{data_ident[0:4]}-{data_ident[4:6]}-{data_ident[6:8]}"
        file_info = prepare_file_info(urljoin(url, file_link), date, level)
        result.append(file_info)
    return result


def cleanup_value(value: str) -> str:
    """Cleanup value using regex patterns."""
    return SPACES.sub(" ", PUNCTUATION.sub("", QUOTES.sub(" ", str(value)))).strip()


def format_employee_name(value: str) -> str:
    """Format employee name."""
    return " ".join([part.capitalize() for part in value.split()])


def format_org_name(value: str) -> str:
    """Replace full organization name with abbreviation."""
    org_names = {
        "Государствнное": "Государственное",
        "уччреждение": "учреждение",
        "бюджетного": "бюджетное",
        " образовательное": " общеобразовательное",
        "им.": "имени",
        "Государственное бюджетное общеобразовательное учреждение города Москвы": "ГБОУ",
        "Государственное бюджетное общеобразовательное учреждение": "ГБОУ",
        "Государственное казенное общеобразовательное учреждение города Москвы": "ГКОУ",
        "Государственное автономное общеобразовательное учреждение города Москвы": "ГАОУ",
        "Государственное автономное общеобразовательное учреждение дополнительного профессионального образования города Москвы": "ГАОУ ДПО",
        "Государственное автономное профессиональное общеобразовательное учреждение города Москвы": "ГАПОУ",
        "Государственное бюджетное профессиональное общеобразовательное учреждение города Москвы": "ГБПОУ",
        "Государственное общеобразовательное учреждение города Москвы": "ГОУ",
        "Муниципальное автономное общеобразовательное учреждение": "МАОУ",
        "Федеральное казенное учреждение": "ФКУ",
    }
    for name, abbreviation in org_names.items():
        value = value.replace(name, abbreviation)
    return value


def format_cell(cell_num: int, cell_value: str) -> str:
    """Format cell value."""
    cell_org_name = (2, 6)
    cell_employee_name = (5,)
    result = cleanup_value(cell_value)
    if cell_num in cell_org_name:
        result = format_org_name(result)
    if cell_num in cell_employee_name:
        result = format_employee_name(result)
    return result


def parse_sheet_data(data: tuple, filename: str) -> list:
    """Parse data and output result as list of lists (rows)."""
    exam_date, exam_level, *_ = filename.split("__")
    logger.debug(
        "parse file: %s (date %s, level %s, rows %s)",
        filename,
        exam_date,
        exam_level,
        len(data),
    )

    result = []
    for row in data:
        # process only rows where first cell (row counter) is int (1, 2, 3...) or float (5.333, 8.833...)
        if not isinstance(row[0].value, int | float):
            continue
        parsed_row = [filename, exam_date, exam_level]
        # skip first cell (row counter)
        for cell_num in range(1, len(row)):
            cell = row[cell_num].value
            if cell:
                cell = format_cell(cell_num, cell)
            # append any cell values, including None
            parsed_row.append(cell)
        result.append(parsed_row)
    return result


def load_sheet_data(file_path: Path) -> tuple:
    """Detect target sheet in Excel file and load its data."""
    wb = load_workbook(file_path)
    sheet_mark = "список"
    sheet_count = 0
    sheet_name = ""
    for name in wb.sheetnames:
        ws = wb[name]
        try:
            first_row = next(ws.rows)
        except StopIteration:
            # empty sheet
            continue
        if (first_cell := first_row[0].value) and sheet_mark in str(first_cell).lower():
            sheet_count += 1
            sheet_name = name

    if sheet_count > 1:
        msg = f"{file_path.name}: there are more than 1 sheet with '{sheet_mark}' in first row"
        raise InvalidFileError(msg)

    if not sheet_name:
        msg = f"{file_path.name}: there is no sheet with '{sheet_mark}' in first row"
        raise InvalidFileError(msg)

    ws = wb[sheet_name]
    first_row = next(ws.rows)

    if len(first_row) < 7:  # noqa: PLR2004
        msg = f"{file_path.name}: wrong number of columns"
        raise InvalidFileError(msg)

    header_rows_count = 2
    sheet_data_without_headers = tuple(ws.iter_rows(min_row=1 + header_rows_count))
    return sheet_data_without_headers  # noqa: RET504


def save_to_csv(csv_file: Path) -> None:
    """Save list of rows to CSV in local file."""
    with csv_file.open("w+", newline="", encoding="utf-8") as fp:
        a = csv.writer(fp, delimiter=";")
        a.writerow(COLUMN_HEADERS)

    parent_dir = Path(csv_file).parent
    for exam_file in parent_dir.glob("*.xlsx"):
        if not (sheet_data := load_sheet_data(exam_file)):
            continue
        result = parse_sheet_data(sheet_data, exam_file.name)
        with Path(csv_file).open("a+", newline="", encoding="utf-8") as fp:
            a = csv.writer(fp, delimiter=";")
            for line in result:
                a.writerow(line)


def save_to_stream(path: Path) -> StringIO:
    """Save list of rows to CSV in string buffer (memory file)."""
    stream = StringIO()
    writer = csv.writer(stream, delimiter="\t", quotechar="'")
    writer.writerow(COLUMN_HEADERS)

    for exam_file in path.glob("*.xlsx"):
        if not (sheet_data := load_sheet_data(exam_file)):
            continue
        result = parse_sheet_data(sheet_data, exam_file.name)
        for line in result:
            writer.writerow(line)
    stream.seek(0)
    return stream


def main():
    """Download and process files, then output result to local CSV file."""
    path = Path("data")
    path.mkdir(parents=True, exist_ok=True)
    csv_file = path.joinpath(f"{path}.csv")
    urls = [
        "http://rcoi.mcko.ru/organizers/schedule/oge/?period=1",
        "http://rcoi.mcko.ru/organizers/schedule/ege/?period=1",
        "http://rcoi.mcko.ru/organizers/schedule/oge/?period=2",
        "http://rcoi.mcko.ru/organizers/schedule/ege/?period=2",
        "http://rcoi.mcko.ru/organizers/schedule/oge/?period=3",
        "http://rcoi.mcko.ru/organizers/schedule/ege/?period=3",
    ]
    with ThreadPool(4) as pool:
        files_info = pool.map(get_files_info, urls)
        files_info_flat = [
            (file["url"], file["name"], path) for files in files_info for file in files
        ]
        pool.starmap(download_file, files_info_flat)
    save_to_csv(csv_file)


if __name__ == "__main__":
    logging.basicConfig(
        format="%(asctime)s  [%(name)s:%(lineno)s]  %(levelname)s - %(message)s",
        level=logging.DEBUG,
    )
    main()
